import streamlit as st
import pandas as pd
import math
import numpy as np
import os
import platform
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.font_manager as fm
from io import BytesIO

# --- 엑셀 스타일링 라이브러리 ---
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl 라이브러리가 필요합니다. requirements.txt에 openpyxl을 추가해주세요.")

# ==============================================================================
# 공통 설정 (Streamlit 페이지 설정 및 한글 폰트 완벽 적용)
# ==============================================================================
st.set_page_config(page_title="맞춤형 벽사다리 시스템", layout="wide")

def set_korean_font():
    os_name = platform.system()
    if os_name == "Windows":
        plt.rc('font', family='Malgun Gothic')
    elif os_name == "Darwin":
        plt.rc('font', family='AppleGothic')
    else:
        # ★ 웹 서버(리눅스) 환경 한글 깨짐 방지: 나눔폰트를 찾아 강제로 메모리에 올림
        font_dirs = ['/usr/share/fonts/truetype/nanum', '/usr/share/fonts/nanum']
        font_files = fm.findSystemFonts(fontpaths=font_dirs)
        for font_file in font_files:
            fm.fontManager.addfont(font_file)
        plt.rc('font', family='NanumGothic')
    plt.rcParams['axes.unicode_minus'] = False

def draw_dim_text(ax, x, y, text, angle=0, color='black', fontsize=11.5, bg_alpha=0.85):
    if angle > 90: angle -= 180
    elif angle < -90: angle += 180
    ax.text(x, y, text, color=color, fontsize=fontsize, fontweight='bold', ha='center', va='center', rotation=angle,
            bbox=dict(facecolor='white', alpha=bg_alpha, edgecolor='none', pad=1.5))

def draw_pipe(ax, x1, y1, x2, y2, t, zorder=1, facecolor='white'):
    dx = x2 - x1
    dy = y2 - y1
    length = math.hypot(dx, dy)
    if length == 0: return
    nx = -dy / length * (t / 2)
    ny = dx / length * (t / 2)
    poly = plt.Polygon(
        [[x1+nx, y1+ny], [x2+nx, y2+ny], [x2-nx, y2-ny], [x1-nx, y1-ny]],
        facecolor=facecolor, edgecolor='black', linewidth=1.2, zorder=zorder
    )
    ax.add_patch(poly)

# ==============================================================================
# 엑셀 저장 시스템
# ==============================================================================
def save_ladder_excel(raw_data, total_sets):
    df = pd.DataFrame(raw_data)
    df_grouped = df.groupby(["구분", "품명", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)"]).agg({'1대당 수량': 'sum'}).reset_index()
    
    sort_mapping = {
        "상현대(상단)": 1, "하현대(조각)": 2, 
        "스나기": 3, "수직다대": 4, 
        "살대(일반)": 5
    }
    df_grouped['정렬키'] = df_grouped['구분'].map(sort_mapping).fillna(99)
    df_grouped = df_grouped.sort_values(by=["정렬키", "재단기장(L)"], ascending=[True, False]).drop('정렬키', axis=1)
    
    df_grouped.insert(0, '순번', range(1, len(df_grouped) + 1))
    df_grouped["총 소요 수량"] = ""
    df_grouped["6M 소요본수"] = ""
    
    df_grouped = df_grouped[["순번", "구분", "품명", "1대당 수량", "총 소요 수량", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)", "6M 소요본수"]]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_grouped.to_excel(writer, sheet_name='통합 재단표', index=False, startrow=2)
        ws = writer.sheets['통합 재단표']
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "👉 대각 벽사다리 총 제작 수량 (EA) :"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal="right", vertical="center")
        
        ws['D1'] = total_sets
        ws['D1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['D1'].font = Font(color="FF0000", bold=True, size=14)
        ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
                             top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))
        ws['D1'].border = thin_border

        color_map = {
            "상현대(상단)": PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid"),
            "하현대(조각)": PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid"),
            "스나기": PatternFill(start_color="D2B4DE", end_color="D2B4DE", fill_type="solid"), 
            "수직다대": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
            "살대(일반)": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        }
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 3):
            gubun_val = ws.cell(row=r_idx, column=2).value if r_idx > 3 else None
            for c_idx, cell in enumerate(row, 1):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
                if r_idx == 3:
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    header = ws.cell(row=3, column=c_idx).value
                    if header == "순번": cell.font = Font(bold=True)
                    elif header == "구분":
                        cell.fill = color_map.get(gubun_val, PatternFill(fill_type=None))
                        cell.font = Font(bold=True)
                    elif header == "총 소요 수량":
                        cell.value = f'=$D$1*D{r_idx}' 
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        cell.font = Font(color="0070C0", bold=True)
                    elif header == "6M 소요본수":
                        cell.value = f'=ROUNDUP((E{r_idx}*F{r_idx})/6000, 1)'
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        cell.font = Font(color="0070C0", bold=True)
                    elif header in ["1대당 수량", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)"]:
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="C00000", bold=True)
                    elif r_idx % 2 == 0: cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max([len(str(ws.cell(row=r, column=col_idx).value)) for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=col_idx).value] + [0])
            ws.column_dimensions[col_letter].width = max(max_len + 6, 12)
            
    return output.getvalue()

# ==============================================================================
# 메인 웹 UI
# ==============================================================================
def main():
    set_korean_font()
    st.title("🏗️ 하나천막기업 - 맞춤형 대각 벽사다리 시스템")
    st.markdown("---")

    with st.container():
        st.subheader("📐 1. 기본 치수 및 분할 입력 (cm 단위)")
        col1, col2, col3 = st.columns(3)
        L_cm = col1.number_input("1. 전체 총기장(cm)", value=2000.0, step=10.0)
        H_left_cm = col2.number_input("2. 좌측 외경 높이(cm)", value=100.0, step=10.0)
        H_right_cm = col3.number_input("3. 우측 외경 높이(cm)", value=70.0, step=10.0)

        col4, col5 = st.columns(2)
        n_sec_m = col4.number_input("4. 스나기 구역(큰 칸) 등분 수", value=5, min_value=1, step=1)
        sub_div = col5.number_input("5. 스나기 안쪽 살대 구역 등분 수", value=4, min_value=1, step=1)

        col6, col7 = st.columns(2)
        total_sets = col6.number_input("6. 총 제작 수량(세트)", value=1, min_value=1, step=1)
        offset_mm = col7.number_input("7. 살대 이격 거리(mm)", value=10.0, step=1.0)
        offset_cm = offset_mm / 10.0

    with st.container():
        st.subheader("⚙️ 2. 파이프 규격 설정 (mm 단위)")
        col8, col9, col10, col11 = st.columns(4)
        p_main = col8.number_input("상하현대 파이프(mm)", value=42.2)
        p_snagi = col9.number_input("스나기 파이프(mm)", value=89.1)
        p_v = col10.number_input("다대 파이프(mm)", value=38.1)
        p_diag = col11.number_input("살대 파이프(mm)", value=31.8)

    st.markdown("---")

    if st.button("🚀 도면 및 산출표 생성하기", type="primary"):
        with st.spinner('도면을 생성하는 중입니다...'):
            t_main_cm, t_snagi_cm, t_v_cm, t_diag_cm = p_main/10, p_snagi/10, p_v/10, p_diag/10
            raw_data = []

            m_slope = (H_right_cm - H_left_cm) / L_cm
            theta_rad = math.atan(m_slope)
            t_slope_deg = math.degrees(theta_rad)
            cos_th = math.cos(theta_rad)
            
            def get_y_top_outer(x): return H_left_cm + m_slope * x
            def get_y_top_inner(x): return get_y_top_outer(x) - (t_main_cm / cos_th)
            def get_y_bot_inner(x): return t_main_cm

            top_chord_len = math.hypot(L_cm, H_right_cm - H_left_cm)
            raw_data.append({"구분": "상현대(상단)", "품명": f"{p_main}mm", "재단기장(L)": round(top_chord_len, 1), 
                             "상단 가공각(°)": abs(int(round(t_slope_deg))), "하단 가공각(°)": abs(int(round(t_slope_deg))), "1대당 수량": 1})

            fig, ax = plt.subplots(figsize=(max(30, L_cm/20), 15), dpi=120)
            plt.subplots_adjust(left=0.08, right=0.92, top=0.85, bottom=0.25)

            pts_top = [
                [0, H_left_cm], [L_cm, H_right_cm],
                [L_cm, H_right_cm - t_main_cm/cos_th], [0, H_left_cm - t_main_cm/cos_th]
            ]
            ax.add_patch(plt.Polygon(pts_top, facecolor='#C0C0C0', edgecolor='black', zorder=5))
            
            offset_y = 65
            ax.annotate('', xy=(0, H_left_cm + offset_y), xytext=(L_cm, H_right_cm + offset_y), arrowprops=dict(arrowstyle='<->', color='navy', lw=2))
            ax.text(L_cm/2, (H_left_cm+H_right_cm)/2 + offset_y + 15, f"상현대 대각길이: {top_chord_len:.1f}cm", ha='center', color='navy', weight='bold', fontsize=20, rotation=t_slope_deg)

            total_intervals = n_sec_m * sub_div
            sub_gap = L_cm / total_intervals 

            x_coords = []
            for k in range(total_intervals + 1):
                x = k * sub_gap
                if k == 0: x = t_snagi_cm / 2
                elif k == total_intervals: x = L_cm - t_snagi_cm / 2
                x_coords.append(x)

            for i in range(n_sec_m):
                x1 = t_snagi_cm if i == 0 else x_coords[i * sub_div] + t_snagi_cm / 2
                x2 = L_cm - t_snagi_cm if i == n_sec_m - 1 else x_coords[(i+1) * sub_div] - t_snagi_cm / 2
                bot_piece_len = x2 - x1
                
                draw_pipe(ax, x1, t_main_cm/2, x2, t_main_cm/2, t_main_cm, zorder=3, facecolor='#A6ACAF')
                raw_data.append({"구분": "하현대(조각)", "품명": f"{p_main}mm", "재단기장(L)": round(bot_piece_len, 1), "상단 가공각(°)": 0, "하단 가공각(°)": 0, "1대당 수량": 1})
                
                cx = (x1 + x2) / 2
                ax.annotate('', xy=(x1, -20), xytext=(x2, -20), arrowprops=dict(arrowstyle='<->', color='#C0392B', lw=1.5))
                ax.text(cx, -20, f"하단 재단: {bot_piece_len:.1f}", ha='center', va='center', color='#C0392B', fontsize=14, weight='bold', bbox=dict(facecolor='white', alpha=0.9, edgecolor='none', pad=2))

                d_xs = [x_coords[i * sub_div + j] for j in range(1, sub_div)]
                if d_xs: 
                    pts_marking = [x1] + d_xs + [x2]
                    for m in range(len(pts_marking) - 1):
                        span_val = pts_marking[m+1] - pts_marking[m]
                        span_cx = (pts_marking[m] + pts_marking[m+1]) / 2
                        
                        if m == 0: lbl = "끝~싱"
                        elif m == len(pts_marking) - 2: lbl = "싱~끝"
                        else: lbl = "싱~싱"
                        
                        y_pos = -70 if m % 2 == 0 else -95
                        ax.text(span_cx, y_pos, f"{lbl}\n{span_val:.1f}", ha='center', va='center', fontsize=12, color='black', bbox=dict(facecolor='white', alpha=0.9, edgecolor='gray', boxstyle='round,pad=0.3'))

            for k in range(total_intervals + 1):
                x = x_coords[k]
                is_snagi = (k % sub_div == 0)
                curr_t_cm = t_snagi_cm if is_snagi else t_v_cm
                x_left, x_right = x - curr_t_cm / 2, x + curr_t_cm / 2
                
                y_top_left, y_top_right = get_y_top_inner(x_left), get_y_top_inner(x_right)
                y_top_max = max(y_top_left, y_top_right)
                cut_angle = abs(int(round(t_slope_deg)))
                
                if is_snagi:
                    y_bot = -30
                    pts = [[x_left, y_bot], [x_right, y_bot], [x_right, y_top_right], [x_left, y_top_left]]
                    ax.add_patch(plt.Polygon(pts, facecolor='#8E44AD', edgecolor='black', zorder=2, linewidth=1.2))
                    v_len_snagi = y_top_max - y_bot
                    raw_data.append({"구분": "스나기", "품명": f"{p_snagi}mm", "재단기장(L)": round(v_len_snagi, 1), "상단 가공각(°)": cut_angle, "하단 가공각(°)": 0, "1대당 수량": 1})
                    
                    draw_dim_text(ax, x, -50, f"L:{v_len_snagi:.1f} ({cut_angle}°)", angle=90, color='purple', fontsize=12, bg_alpha=0)
                    if k == 0:
                        ax.annotate('', xy=(x - curr_t_cm, 0), xytext=(x - curr_t_cm, -30), arrowprops=dict(arrowstyle='<->', color='purple', lw=1.5))
                        ax.text(x - curr_t_cm - 2, -15, f"연장 30cm", ha='right', va='center', color='purple', fontsize=14, weight='bold')
                else:
                    y_bot = t_main_cm
                    pts = [[x_left, y_bot], [x_right, y_bot], [x_right, y_top_right], [x_left, y_top_left]]
                    ax.add_patch(plt.Polygon(pts, facecolor='#2980B9', edgecolor='black', zorder=2, linewidth=1.2))
                    v_len_dadae = y_top_max - y_bot
                    raw_data.append({"구분": "수직다대", "품명": f"{p_v}mm", "재단기장(L)": round(v_len_dadae, 1), "상단 가공각(°)": cut_angle, "하단 가공각(°)": 0, "1대당 수량": 1})
                    
                    draw_dim_text(ax, x, y_top_max + 35, f"L:{v_len_dadae:.1f} ({cut_angle}°)", angle=90, color='black', fontsize=12, bg_alpha=1.0)

            is_forward = (H_left_cm >= H_right_cm) 
            for k in range(total_intervals):
                x_L, x_R = x_coords[k], x_coords[k+1]
                r_L = (t_snagi_cm if k % sub_div == 0 else t_v_cm) / 2
                r_R = (t_snagi_cm if (k+1) % sub_div == 0 else t_v_cm) / 2
                
                wx_start = x_L + r_L + offset_cm
                wx_end = x_R - r_R - offset_cm
                if wx_end - wx_start <= 0: continue

                mid_x = (wx_start + wx_end) / 2
                v_len_est = abs(get_y_top_inner(mid_x) - get_y_bot_inner(mid_x))
                diag_len_est = math.hypot(wx_end - wx_start, v_len_est)
                sin_theta = v_len_est / diag_len_est if diag_len_est > 0 else 1
                w_half = (t_diag_cm / 2) / sin_theta if sin_theta > 0.01 else t_diag_cm / 2

                px_bot = wx_start + w_half if is_forward else wx_end - w_half
                px_top = wx_end - w_half if is_forward else wx_start + w_half

                if is_forward:
                    p_bl, p_br, p_tr, p_tl = [px_bot - w_half, get_y_bot_inner(px_bot - w_half)], [px_bot + w_half, get_y_bot_inner(px_bot + w_half)], [px_top + w_half, get_y_top_inner(px_top + w_half)], [px_top - w_half, get_y_top_inner(px_top - w_half)]
                    pts = [p_bl, p_br, p_tr, p_tl]
                else:
                    p_br, p_bl, p_tl, p_tr = [px_bot + w_half, get_y_bot_inner(px_bot + w_half)], [px_bot - w_half, get_y_bot_inner(px_bot - w_half)], [px_top - w_half, get_y_top_inner(px_top - w_half)], [px_top + w_half, get_y_top_inner(px_top + w_half)]
                    pts = [p_br, p_bl, p_tl, p_tr]

                poly = plt.Polygon(pts, facecolor='#F1C40F', edgecolor='black', linewidth=1.2, zorder=1)
                ax.add_patch(poly)

                py_bot, py_top = get_y_bot_inner(px_bot), get_y_top_inner(px_top)
                diag_l_raw = math.hypot(px_top - px_bot, py_top - py_bot)
                ux = (px_top - px_bot) / diag_l_raw if diag_l_raw > 0 else 0
                uy = (py_top - py_bot) / diag_l_raw if diag_l_raw > 0 else 1
                projections = [(pt[0] - px_bot) * ux + (pt[1] - py_bot) * uy for pt in pts]
                max_diag_len = max(projections) - min(projections)

                angle_D = math.atan2(py_top - py_bot, px_top - px_bot)
                angle_diff_top = abs(angle_D - math.atan(m_slope)) % math.pi
                if angle_diff_top > math.pi/2: angle_diff_top = math.pi - angle_diff_top
                d_top_angle = int(round(math.degrees(math.pi/2 - angle_diff_top)))

                angle_diff_bot = abs(angle_D - 0) % math.pi
                if angle_diff_bot > math.pi/2: angle_diff_bot = math.pi - angle_diff_bot
                d_bot_angle = int(round(math.degrees(math.pi/2 - angle_diff_bot)))

                mx, my = (px_bot + px_top) / 2, (py_bot + py_top) / 2
                draw_dim_text(ax, mx, my, f"L:{max_diag_len:.1f} ({d_top_angle}°/{d_bot_angle}°)", angle=math.degrees(angle_D), color='#8B0000', fontsize=12, bg_alpha=1.0)

                raw_data.append({"구분": "살대(일반)", "품명": f"{p_diag}mm", "재단기장(L)": round(max_diag_len, 1), "상단 가공각(°)": d_top_angle, "하단 가공각(°)": d_bot_angle, "1대당 수량": 1})

            max_h = max(H_left_cm, H_right_cm)
            ax.annotate('', xy=(0, -125), xytext=(L_cm, -125), arrowprops=dict(arrowstyle='<->', color='red', lw=2))
            ax.text(L_cm/2, -125, f"전체 총기장: {L_cm}cm", ha='center', va='center', color='red', weight='bold', fontsize=26, bbox=dict(facecolor='white', alpha=0.9, edgecolor='none', pad=2))
            
            ax.annotate('', xy=(-20, 0), xytext=(-20, H_left_cm), arrowprops=dict(arrowstyle='<->', color='black', lw=1.5))
            ax.text(-25, H_left_cm/2, f"좌측 외경\n{H_left_cm}cm", ha='right', va='center', color='black', weight='bold', fontsize=18)
            
            ax.annotate('', xy=(L_cm + 20, 0), xytext=(L_cm + 20, H_right_cm), arrowprops=dict(arrowstyle='<->', color='black', lw=1.5))
            ax.text(L_cm + 25, H_right_cm/2, f"우측 외경\n{H_right_cm}cm", ha='left', va='center', color='black', weight='bold', fontsize=18)

            ax.set_title(f"■ 맞춤형 대각 벽사다리 ■", fontsize=36, fontweight='bold', pad=45)
            ax.set_xlim(-max(70, L_cm*0.08), L_cm + max(70, L_cm*0.08))
            ax.set_ylim(-190, max_h + 180)
            ax.axis('off')
            ax.set_aspect('equal')

            st.pyplot(fig)

            pdf_buffer = BytesIO()
            fig.savefig(pdf_buffer, format="pdf", bbox_inches='tight')
            pdf_buffer.seek(0)

            excel_data = save_ladder_excel(raw_data, total_sets)

            st.success("✅ 도면 및 산출표 생성이 완료되었습니다!")
            col_btn1, col_btn2 = st.columns(2)
            
            with col_btn1:
                st.download_button(
                    label="📥 PDF 도면 다운로드",
                    data=pdf_buffer,
                    file_name=f"벽사다리_L{int(L_cm)}_H{int(H_left_cm)}x{int(H_right_cm)}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            with col_btn2:
                st.download_button(
                    label="📊 엑셀 산출표 다운로드",
                    data=excel_data,
                    file_name=f"벽사다리_산출표_L{int(L_cm)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

if __name__ == "__main__":
    main()
